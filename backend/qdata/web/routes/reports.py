import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, DoughnutChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.chart.label import DataLabelList
from datetime import date, datetime
from pydantic import BaseModel
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from sqlalchemy.orm import joinedload, defer

from qdata.auth.dependencies import get_current_user
from qdata.auth.permissions import require_role
from qdata.core.descriptions import describe_detail, describe_error, duplicate_groups, describe_rule_simple, GLOSARIO, QUE_HACER, RULE_DISPLAY_NAMES, severity_description
from qdata.db.models import ErrorAction, GroupPermission, Project, Report, User
from qdata.db.session import get_session
from qdata.rules.base import (
    MAX_DUPE_ROWS_PER_GROUP,
    MAX_SAMPLE_FAILURES,
    MAX_SAMPLE_FAILURES_GENERIC,
    PERSONAS_SIMILARES_RULES,
)


class SetActionRequest(BaseModel):
    status: str

router = APIRouter()


@router.get("/")
@router.get("")
async def list_reports(
    group_id: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
    limit: int = 20,
    offset: int = 0,
):
    if user.role == "admin":
        query = select(Report).options(
            joinedload(Report.project),
            defer(Report.result_json),
            defer(Report.recommendations),
        )
    else:
        subq = select(GroupPermission.group_id).where(GroupPermission.user_id == user.id)
        query = (
            select(Report)
            .options(
                joinedload(Report.project),
                defer(Report.result_json),
                defer(Report.recommendations),
            )
            .where(
                or_(
                    Report.user_id == user.id,
                    Report.project.has(Project.group_id.in_(subq)),
                )
            )
        )
    if group_id:
        query = query.where(Report.project.has(group_id=group_id))
    if start_date:
        query = query.where(Report.executed_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.where(Report.executed_at <= datetime.combine(end_date, datetime.max.time()))
    result = await session.execute(
        query.order_by(Report.executed_at.desc()).offset(offset).limit(limit)
    )
    reports = result.unique().scalars().all()

    def _extract_names(sc: dict | None):
        if not sc:
            return None, None, None
        st = sc.get("source_type") or ""
        cs = sc.get("connection_string") or ""
        fp = sc.get("file_path") or ""
        query = sc.get("query") or ""
        import re as _re
        table_name = None
        if query:
            m = _re.search(r"FROM\s+[`\"']?(\w+)[`\"']?", query, _re.IGNORECASE)
            if m:
                table_name = m.group(1).upper()
        if st in ("mysql", "postgresql", "sqlite", "mssql"):
            source_label = table_name or st.upper()
            db_name = cs.rsplit("/", 1)[-1].split("?")[0] if "/" in cs else None
            connection_label = db_name.upper() if db_name else cs
        elif st == "file":
            source_label = "Archivo"
            connection_label = fp.rsplit("/", 1)[-1].rsplit("\\", 1)[-1] if fp else None
        else:
            source_label = st or "Desconocido"
            m = _re.search(r"(?:DATABASE|Database)=([^;]+)", cs)
            db_name = m.group(1).upper() if m else None
            if not db_name:
                db_name = cs.rsplit("/", 1)[-1].split("?")[0] if "/" in cs else None
            connection_label = db_name or cs or fp or None
        return source_label, connection_label, st

    return [
        {
            "id": str(r.id),
            "project_id": str(r.project_id),
            "project_name": r.project.name if r.project else None,
            "source_type": _extract_names(r.project.source_config if r.project else None)[2],
            "source_label": _extract_names(r.project.source_config if r.project else None)[0],
            "connection_label": _extract_names(r.project.source_config if r.project else None)[1],
            "score": r.score,
            "label": r.label,
            "summary": r.summary,
            "executed_at": r.executed_at.isoformat() if r.executed_at else None,
        }
        for r in reports
    ]


@router.get("/{report_id}")
async def get_report(
    report_id: str,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if user.role == "admin":
        base_q = select(Report).options(joinedload(Report.project)).where(Report.id == report_id)
    else:
        subq = select(GroupPermission.group_id).where(GroupPermission.user_id == user.id)
        base_q = (
            select(Report)
            .options(joinedload(Report.project))
            .where(
                Report.id == report_id,
                or_(
                    Report.user_id == user.id,
                    Report.project.has(Project.group_id.in_(subq)),
                )
            )
        )
    result = await session.execute(base_q)
    report = result.unique().scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    source_query = None
    selected_columns: list = []
    if report.project and report.project.source_config:
        source_query = report.project.source_config.get("query")
        selected_columns = report.project.source_config.get("selected_columns") or []
    result = report.result_json or {}
    if isinstance(result, dict):
        for r in result.get("results") or []:
            if not isinstance(r, dict):
                continue
            sfs = r.get("sample_failures")
            if not isinstance(sfs, list):
                continue
            if r.get("rule_name") == "duplicate_check":
                legacy = any(isinstance(it, dict) and isinstance(it.get("rows"), list) for it in sfs)
                if legacy:
                    for item in sfs:
                        if isinstance(item, dict):
                            rows = item.get("rows")
                            if isinstance(rows, list):
                                item["size"] = len(rows)
                    sfs.sort(key=lambda it: (it.get("size") or 0) if isinstance(it, dict) else 0, reverse=True)
                    for item in sfs[:MAX_SAMPLE_FAILURES]:
                        if isinstance(item, dict):
                            rows = item.get("rows")
                            if isinstance(rows, list):
                                item["rows"] = rows[:MAX_DUPE_ROWS_PER_GROUP]
                    r["sample_failures"] = sfs[:MAX_SAMPLE_FAILURES]
            elif r.get("rule_name") not in PERSONAS_SIMILARES_RULES and len(sfs) > MAX_SAMPLE_FAILURES_GENERIC:
                r["sample_failures"] = sfs[:MAX_SAMPLE_FAILURES_GENERIC]
    return {
        "id": str(report.id),
        "project_id": str(report.project_id) if report.project_id else None,
        "project_name": report.project.name if report.project else None,
        "source_query": source_query,
        "selected_columns": selected_columns,
        "score": report.score,
        "label": report.label,
        "result": result,
        "recommendations": report.recommendations,
        "summary": report.summary,
        "executed_at": report.executed_at.isoformat() if report.executed_at else None,
    }


@router.delete("/{report_id}")
async def delete_report(
    report_id: str,
    user: User = Depends(require_role(["admin"])),
    session: AsyncSession = Depends(get_session),
):
    result = await session.execute(select(Report).where(Report.id == report_id))
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    await session.delete(report)
    await session.commit()
    return {"status": "deleted"}


@router.get("/{report_id}/actions")
async def get_report_actions(
    report_id: str,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    result = await session.execute(
        select(ErrorAction).where(ErrorAction.report_id == report_id)
    )
    actions = result.scalars().all()
    return [
        {
            "error_index": a.error_index,
            "rule_index": a.rule_index,
            "status": a.status,
            "updated_at": a.updated_at.isoformat() if a.updated_at else None,
        }
        for a in actions
    ]


@router.get("/{report_id}/rules/{rule_idx}/actions")
async def get_rule_actions(
    report_id: str,
    rule_idx: int,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    result = await session.execute(
        select(ErrorAction).where(
            ErrorAction.report_id == report_id,
            ErrorAction.rule_index == rule_idx,
        )
    )
    actions = result.scalars().all()
    return [
        {
            "error_index": a.error_index,
            "status": a.status,
            "updated_at": a.updated_at.isoformat() if a.updated_at else None,
        }
        for a in actions
    ]


@router.put("/{report_id}/rules/{rule_idx}/errors/{error_idx}/action")
async def set_error_action(
    report_id: str,
    rule_idx: int,
    error_idx: int,
    body: SetActionRequest,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if body.status not in ("sin_accion", "en_revision", "solucionado"):
        raise HTTPException(status_code=400, detail="Invalid status")
    result = await session.execute(
        select(ErrorAction).where(
            ErrorAction.report_id == report_id,
            ErrorAction.rule_index == rule_idx,
            ErrorAction.error_index == error_idx,
        )
    )
    existing = result.scalar_one_or_none()
    if existing:
        existing.status = body.status
        existing.updated_at = datetime.utcnow()
    else:
        action = ErrorAction(
            report_id=report_id,
            rule_index=rule_idx,
            error_index=error_idx,
            status=body.status,
        )
        session.add(action)
    await session.commit()
    return {"status": body.status}


def _build_summary(results: list[dict], score: int, project_name: str | None = None) -> str:
    total_rules = len(results)
    total_records = max((r.get("total", 0) for r in results), default=0)
    total_errors = sum(r.get("failed", 0) for r in results)
    failed_rules = [r for r in results if not r.get("pass", r.get("passed", False))]
    findings = []
    for r in sorted(failed_rules, key=lambda x: x.get("failed", 0), reverse=True)[:3]:
        dname = RULE_DISPLAY_NAMES.get(r.get("rule_name", ""), r.get("rule_name", ""))
        findings.append(f"{dname} ({r.get('failed', 0):,} errores)")

    project_part = f" del proyecto '{project_name}'" if project_name else ""
    lines = [
        f"Este reporte revisó {total_records:,} registros{project_part}. "
        f"Encontramos {total_errors:,} errores en total distribuidos en {len(failed_rules)} "
        f"de {total_rules} reglas aplicadas."
    ]
    if findings:
        lines.append(f"Los principales problemas: {', '.join(findings)}.")
    if total_records > 0:
        if score >= 80:
            lines.append("Esto significa que la mayoría de los registros pasaron todas las validaciones.")
        elif score >= 50:
            lines.append("Esto significa que aproximadamente la mitad de los registros pasaron todas las validaciones.")
        else:
            lines.append("Esto significa que casi ningún registro pasó todas las validaciones (algunos pueden fallar en más de una regla).")
        lines.append("Se recomienda revisar los datos antes de usarlos para análisis o reportes.")
    return " ".join(lines)


def _severity_unit(rule_name: str) -> str:
    rl = rule_name.lower()
    if "email" in rl or "correo" in rl:
        return "correos"
    if "phone" in rl or "telefono" in rl or "teléfono" in rl:
        return "teléfonos"
    if "duplicate" in rl or "duplicado" in rl:
        return "grupos"
    return "registros"


def _severity_text(failure_pct: float, total: int, failed: int, rule_name: str = "") -> str:
    if failed == 0 or total == 0:
        return "Sin errores"
    actual_pct = round(failed / total * 100, 1)
    unit = _severity_unit(rule_name)
    level = "La mayoría" if failure_pct >= 50 else "Muchos" if failure_pct >= 30 else "Algunos" if failure_pct >= 10 else "Pocos"
    return f"{level} errores — {actual_pct}% de los {unit}"


def _severity_fill(severity: str) -> PatternFill:
    m = {"error": "FEE2E2", "warning": "FEF3C7", "info": "DBEAFE"}
    return PatternFill(start_color=m.get(severity, "F3F4F6"), end_color=m.get(severity, "F3F4F6"), fill_type="solid")


def _pass_fill(passed: bool) -> PatternFill:
    return PatternFill(start_color="DCFCE7" if passed else "FEE2E2", end_color="DCFCE7" if passed else "FEE2E2", fill_type="solid")


_GREEN = "22C55E"
_YELLOW = "EAB308"
_RED = "EF4444"


def _score_color(score: int) -> str:
    return _GREEN if score >= 70 else _YELLOW if score >= 50 else _RED


@router.get("/{report_id}/export/excel")
async def export_report_excel(
    report_id: str,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if user.role == "admin":
        base_q = select(Report).options(joinedload(Report.project)).where(Report.id == report_id)
    else:
        subq = select(GroupPermission.group_id).where(GroupPermission.user_id == user.id)
        base_q = (
            select(Report)
            .options(joinedload(Report.project))
            .where(
                Report.id == report_id,
                or_(
                    Report.user_id == user.id,
                    Report.project.has(Project.group_id.in_(subq)),
                )
            )
        )
    result = await session.execute(base_q)
    report = result.unique().scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    data = report.result_json or {}
    results = data.get("results", [])
    score = report.score or 0
    label = report.label or "N/A"
    summary_text = report.summary or ""

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ============================
    # Sheet 1: Resumen Ejecutivo
    # ============================
    ws = wb.active
    ws.title = "Resumen Ejecutivo"

    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, "REPORTE DE CALIDAD DE DATOS")
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    c.alignment = Alignment(horizontal="center")

    # Score display
    sc = _score_color(score)
    ws.merge_cells("A2:H2")
    c = ws.cell(2, 1, f"Puntaje: {score}/100")
    c.font = Font(bold=True, size=28, color="FFFFFF")
    c.fill = PatternFill(start_color=sc, end_color=sc, fill_type="solid")
    c.alignment = Alignment(horizontal="center")

    # Label
    ws.merge_cells("A3:H3")
    c = ws.cell(3, 1, f"{label.upper()}")
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill(start_color=sc, end_color=sc, fill_type="solid")
    c.alignment = Alignment(horizontal="center")

    # Summary
    ws.merge_cells("A5:G5")
    project_name = report.project.name if report.project else None
    gen_summary = _build_summary(results, score, project_name)
    c = ws.cell(5, 1, gen_summary)
    c.font = Font(size=11, color="333333")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[5].height = 55

    # Rules table header
    rstart = 7
    cols_mgmt = ["", "Regla", "Qué hace", "Severidad", "Estado", "¿Qué tan grave es?", "Acción recomendada", ""]
    for ci, h in enumerate(cols_mgmt, 1):
        c = ws.cell(rstart, ci, h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for ri, r in enumerate(results, rstart + 1):
        rname = r.get("rule_name", "")
        dname = RULE_DISPLAY_NAMES.get(rname, rname)
        passed = r.get("pass", r.get("passed", False))
        severity = r.get("severity", "info")
        failed = r.get("failed", 0)
        total = r.get("total", 0)
        pct = r.get("failure_pct", 0)

        # Traffic-light bar column
        sc = _score_color(100 - pct)
        bar = ws.cell(ri, 1, "")
        bar.fill = PatternFill(start_color=sc, end_color=sc, fill_type="solid")
        bar.border = thin_border

        ws.cell(ri, 2, dname).border = thin_border
        ws.cell(ri, 3, describe_rule_simple(rname)).border = thin_border
        c = ws.cell(ri, 4, severity.upper())
        c.fill = _severity_fill(severity)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

        estado = "Bien" if passed else "Revisar"
        c = ws.cell(ri, 5, estado)
        c.fill = _pass_fill(passed)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

        c = ws.cell(ri, 6, _severity_text(pct, total, failed, rname))
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True)
        if not passed:
            c.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        c = ws.cell(ri, 7, r.get("recommendation") or "")
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True)

        # Column H: progress bar value (DataBar applied below)
        total_r = r.get("total", 0)
        failed_r = r.get("failed", 0)
        pv = round(failed_r / total_r * 100, 1) if total_r > 0 else 0
        ws.cell(ri, 8, pv).border = thin_border

    # DataBar progress bar on column H
    if len(results) > 0:
        rule = DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=100,
            color="6366F1",
            showValue=None,
        )
        ws.conditional_formatting.add(f"H{rstart+1}:H{rstart+len(results)}", rule)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 32
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 0.5

    # Bar chart — Fallos por regla
    last_data_row = rstart + len(results)
    if len(results) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Fallos por regla"
        chart.y_axis.title = "Cantidad de errores"
        chart.style = 10
        chart.width = 22
        chart.height = 14

        fail_col = 10  # column J
        ws.cell(rstart, fail_col, "Fallos")
        for ri, r in enumerate(results, rstart + 1):
            ws.cell(ri, fail_col, r.get("failed", 0))

        data_ref = Reference(ws, min_col=fail_col, min_row=rstart, max_row=last_data_row)
        cats_ref = Reference(ws, min_col=2, min_row=rstart + 1, max_row=last_data_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        s = chart.series[0]
        s.graphicalProperties.solidFill = "6366F1"
        ws.add_chart(chart, f"B{last_data_row + 3}")

        # Doughnut chart — Correctos vs Con problemas
        total_rec = max((rr.get("total", 0) for rr in results), default=0)
        total_err = sum(rr.get("failed", 0) for rr in results)
        clean = max(0, total_rec - min(total_err, total_rec))

        doughnut_start = last_data_row + 20
        ws.cell(doughnut_start, 2, "Correctos").font = Font(size=9, color="888888")
        ws.cell(doughnut_start, 3, clean)
        ws.cell(doughnut_start + 1, 2, "Con problemas").font = Font(size=9, color="888888")
        ws.cell(doughnut_start + 1, 3, min(total_err, total_rec))

        dchart = DoughnutChart()
        dchart.title = "Correctos vs Con problemas"
        dchart.style = 10
        dchart.width = 16
        dchart.height = 12

        ddata = Reference(ws, min_col=3, min_row=doughnut_start, max_row=doughnut_start + 1)
        dcats = Reference(ws, min_col=2, min_row=doughnut_start, max_row=doughnut_start + 1)
        dchart.add_data(ddata, titles_from_data=False)
        dchart.set_categories(dcats)
        # Color slices via GraphicialProperties on each DataPoint
        from openpyxl.chart.series import DataPoint
        s0 = DataPoint(idx=0)
        s0.graphicalProperties.solidFill = "22C55E"
        s1 = DataPoint(idx=1)
        s1.graphicalProperties.solidFill = "EF4444"
        dchart.series[0].data_points = [s0, s1]
        ws.add_chart(dchart, f"B{doughnut_start}")

    # ============================
    # Sheet 2: Glosario
    # ============================
    ws2 = wb.create_sheet("Glosario")
    rule_names_in_report = {r.get("rule_name", "") for r in results}
    ws2.cell(1, 1, "Regla").font = header_font
    ws2.cell(1, 1).fill = header_fill
    ws2.cell(1, 1).border = thin_border
    ws2.cell(1, 2, "¿Qué hace?").font = header_font
    ws2.cell(1, 2).fill = header_fill
    ws2.cell(1, 2).border = thin_border
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 70
    gi = 2
    for rname in sorted(rule_names_in_report):
        if rname in GLOSARIO:
            ws2.cell(gi, 1, rname).border = thin_border
            ws2.cell(gi, 2, GLOSARIO[rname]).border = thin_border
            gi += 1

    # ============================
    # Sheet 3: Detalle Técnico
    # ============================
    ws3 = wb.create_sheet("Detalle Técnico")
    row_idx = 1

    def _clean_values(vals: dict) -> dict:
        return {k: v for k, v in vals.items() if not k.startswith("Unnamed:")}

    def _should_skip_columna(items: list, rname: str, rec: str | None) -> bool:
        for it in items:
            info = describe_error(rname, it, rec)
            if info.get("columna") and info["columna"] not in ("—", "-", None, ""):
                return False
        return True

    MAX_DISPLAYED_GROUPS = 15

    for r in results:
        rname = r.get("rule_name", "desconocida")
        dname = RULE_DISPLAY_NAMES.get(rname, rname)
        recommendation = r.get("recommendation")
        details_list = r.get("details", [])
        sample_failures = r.get("sample_failures", [])

        total_failed = r.get("failed", 0)
        total_items = r.get("total", 0)
        failure_pct = r.get("failure_pct", 0)

        passed = r.get("pass", r.get("passed", False))
        hdr_color = "22C55E" if passed else "EF4444"

        base_col_count = 6  # #, Fila, Valor, Descripción, Sugerencia, Qué hacer
        skip_col = _should_skip_columna(sample_failures, rname, recommendation)
        detail_col_count = max(7, base_col_count + (0 if skip_col else 1))

        ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=detail_col_count)
        hdr_text = f"{dname} ({rname}) — {total_failed:,} errores de {total_items:,} registros ({failure_pct:.2f}%)"
        c = ws3.cell(row_idx, 1, hdr_text)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill(start_color=hdr_color, end_color=hdr_color, fill_type="solid")
        c.alignment = Alignment(wrap_text=True)
        row_idx += 1

        # Description line
        if r.get("description"):
            ws3.cell(row_idx, 1, r["description"]).font = Font(italic=True, size=9, color="555555")
            row_idx += 1

        if recommendation:
            ws3.cell(row_idx, 1, f"Recomendación: {recommendation}").font = Font(size=9, color="B45309")
            ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=detail_col_count)
            row_idx += 1

        # Summary sub-section
        if details_list:
            c = ws3.cell(row_idx, 1, "Resumen por columna:")
            c.font = Font(bold=True, size=10, color="6366F1")
            row_idx += 1
            for d in details_list:
                ws3.cell(row_idx, 1, describe_detail(rname, d))
                row_idx += 1

        if not sample_failures:
            ws3.cell(row_idx, 1, "Sin errores de muestra").font = Font(italic=True, color="888888")
            row_idx += 1
            row_idx += 1
            continue

        # --- Grouped display for duplicate_check ---
        if rname == "duplicate_check":
            groups = duplicate_groups(r)
            groups.sort(key=lambda g: g["size"], reverse=True)

            total_groups = len(groups)
            displayed = groups[:MAX_DISPLAYED_GROUPS]
            hidden_count = total_groups - len(displayed)
            hidden_rows = sum(g["size"] for g in groups[MAX_DISPLAYED_GROUPS:])

            note = f"Mostrando los {len(displayed)} grupos más grandes de {total_groups} grupos totales"
            if hidden_count > 0:
                note += f" (+{hidden_count} grupos con {hidden_rows} filas no mostrados)"
            ws3.cell(row_idx, 1, note).font = Font(italic=True, size=9, color="888888")
            row_idx += 1

            for ci, h in enumerate(["#", "Tamaño", "Valores", "Descripción", "Qué hacer"], 1):
                c = ws3.cell(row_idx, ci, h)
                c.font = header_font
                c.fill = header_fill
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
            row_idx += 1

            for gi, g in enumerate(displayed, 1):
                item = g
                info = describe_error(rname, item, recommendation)
                first = (g["rows"] or [{}])[0]
                vals = first.get("values") or {}
                vals_clean = _clean_values(vals)
                parts = []
                has_empty = False
                for k, v in list(vals_clean.items())[:4]:
                    v_display = str(v) if v and str(v).strip() else "Sin dato registrado"
                    if not v or str(v).strip() == "":
                        has_empty = True
                    parts.append(f"{k}={v_display}")
                val_str = ", ".join(parts)

                ws3.cell(row_idx, 1, gi).border = thin_border
                ws3.cell(row_idx, 2, f"{g['size']} filas").border = thin_border
                ws3.cell(row_idx, 3, val_str).border = thin_border
                desc = info.get("descripcion") or ""
                if has_empty:
                    key_label = next(iter(vals_clean.keys()), "dato")
                    desc += f" — Estas {g['size']} filas no tienen {key_label} cargado, no es que estén repetidas con el mismo valor."
                ws3.cell(row_idx, 4, desc).border = thin_border
                ws3.cell(row_idx, 5, info.get("que_hacer") or "").border = thin_border
                row_idx += 1

            if hidden_count > 0:
                ws3.cell(row_idx, 1, f"... y {hidden_count} grupos más ({hidden_rows} filas). Revisa la fuente de datos original para el detalle completo.").font = Font(italic=True, size=9, color="888888")
                ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=detail_col_count)
                row_idx += 1

            row_idx += 1
            continue

        # --- Standard error table for other rules ---
        n_sample = len(sample_failures)
        note = f"Mostrando {n_sample:,} de {total_failed:,} errores" if total_failed > n_sample else f"{n_sample:,} errores"
        c = ws3.cell(row_idx, 1, note)
        c.font = Font(italic=True, size=9, color="888888")
        row_idx += 1

        # Build headers dynamically
        detail_headers = ["#", "Fila"]
        if not skip_col:
            detail_headers.append("Columna")
        detail_headers += ["Valor", "Descripción", "Sugerencia", "Qué hacer"]

        for ci, h in enumerate(detail_headers, 1):
            c = ws3.cell(row_idx, ci, h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
        row_idx += 1

        error_counter = 0
        group_to_num = {}

        for item in sample_failures:
            info = describe_error(rname, item, recommendation)
            group_idx = item.get("group_idx")
            if group_idx is not None:
                if group_idx not in group_to_num:
                    error_counter += 1
                    group_to_num[group_idx] = error_counter
                current_err_num = group_to_num[group_idx]
            else:
                error_counter += 1
                current_err_num = error_counter

            vals_raw = item.get("values", {})
            vals_clean = _clean_values(vals_raw)
            col_idx = 1

            ws3.cell(row_idx, col_idx, current_err_num).border = thin_border; col_idx += 1
            ws3.cell(row_idx, col_idx, info.get("fila") or "—").border = thin_border; col_idx += 1
            if not skip_col:
                ws3.cell(row_idx, col_idx, info.get("columna") or "—").border = thin_border; col_idx += 1
            ws3.cell(row_idx, col_idx, info.get("valor") or "—").border = thin_border; col_idx += 1
            ws3.cell(row_idx, col_idx, info.get("descripcion") or "").border = thin_border; col_idx += 1
            ws3.cell(row_idx, col_idx, info.get("sugerencia") or "").border = thin_border; col_idx += 1
            ws3.cell(row_idx, col_idx, info.get("que_hacer") or "").border = thin_border
            row_idx += 1

        row_idx += 1  # blank row

    for col_letter, w in {"A": 6, "B": 10, "C": 16, "D": 26, "E": 45, "F": 35, "G": 45}.items():
        try:
            ws3.column_dimensions[col_letter].width = w
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_{report_id[:8]}.xlsx"},
    )


@router.get("/{report_id}/export/pdf")
async def export_report_pdf(
    report_id: str,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if user.role == "admin":
        base_q = select(Report).where(Report.id == report_id)
    else:
        subq = select(GroupPermission.group_id).where(GroupPermission.user_id == user.id)
        base_q = (
            select(Report)
            .options(joinedload(Report.project))
            .where(
                Report.id == report_id,
                or_(
                    Report.user_id == user.id,
                    Report.project.has(Project.group_id.in_(subq)),
                )
            )
        )
    result = await session.execute(base_q)
    report = result.unique().scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    data = report.result_json or {}
    results = data.get("results", [])
    recommendations = data.get("recommendations", [])

    from qdata.core.reporter import generate_pdf

    try:
        pdf_bytes = generate_pdf(
            results=results,
            score=report.score or 0,
            label=report.label or "N/A",
            recommendations=recommendations,
            summary=report.summary or "",
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {e}")

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_{report_id[:8]}.pdf"},
    )
